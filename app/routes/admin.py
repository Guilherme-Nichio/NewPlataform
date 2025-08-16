from flask import render_template, request, redirect, session
from werkzeug.security import generate_password_hash
import sqlite3
from datetime import datetime, timedelta
from openpyxl import load_workbook
from conexaoabanco import admin
import matplotlib.pyplot as plt
import io
import base64
from collections import Counter
from itertools import product
def criar_admin():
    with sqlite3.connect('db.sqlite3') as conn:
        c = conn.cursor()
        admin = c.execute("SELECT * FROM usuarios WHERE tipo='admin'").fetchone()
        if not admin:
            senha_hash = generate_password_hash("admin123")  # senha padrão
            c.execute(
                    "INSERT INTO usuarios (email,nome, senha, telefone, tipo) VALUES (?, ?, ?, ?, ?)",
                    ("admin@admin.com","admin", senha_hash, "00000000", "admin")
            )
            conn.commit()
def gerar_contagem_tipos():
    # Pega todas as respostas
    with sqlite3.connect('db.sqlite3') as conn:
        rows = conn.execute("SELECT OxD, SxR, PxN, WxT FROM respostas").fetchall()
    
    tipos_respostas = [''.join(r) for r in rows]
    contagem = Counter(tipos_respostas)

    # Todos os 16 tipos possíveis
    oxd = ['O','D']
    sxr = ['S','R']
    pxn = ['P','N']
    wxt = ['W','T']
    todos_tipos = [''.join(p) for p in product(oxd, sxr, pxn, wxt)]

    # Garantir que todos os 16 tipos apareçam
    resultado = {t: contagem.get(t,0) for t in todos_tipos}
    return resultado    
def register_admin_routes(app):

    def enviar_email(destinatario, senha):
        mensagem = f"Olá! Seu acesso foi criado.\nLogin: {destinatario}\nSenha: {senha}"
        print(">> Email enviado para:", destinatario)
        print(mensagem)
        # Aqui você colocaria o envio real com SMTP ou outro serviço



    @app.route('/admin')
    def admin():
        if 'user_id' not in session:
            return redirect('/login')
        
        with sqlite3.connect('db.sqlite3') as conn:
            total_respostas = conn.execute("SELECT COUNT(*) FROM respostas").fetchone()[0]
            status = conn.execute("""
                SELECT 
                    SUM(CASE WHEN ativo = 1 THEN 1 ELSE 0 END) as ativos,
                    SUM(CASE WHEN ativo = 0 THEN 1 ELSE 0 END) as inativos
                FROM usuarios
            """).fetchone()

        tipos = gerar_contagem_tipos()
        
        return render_template('admin.html', 
                            ativos=status[0] or 0,
                            inativos=status[1] or 0, 
                            qntResposta=total_respostas,
                            tipos=tipos)



    @app.route('/admin/cadastrar', methods=['GET', 'POST'])
    def cadastrar_usuario():
        if session.get('tipo') != 'admin':
            return redirect('/login')

        if request.method == 'POST':
            telefone = request.form['telefone']
            email = request.form['email']
            nome = request.form['nome']
            senha_gerada = 'ENES2025'
            senha_hash = generate_password_hash(senha_gerada)
            validade = (datetime.now() + timedelta(days=365)).strftime('%Y-%m-%d')

            with sqlite3.connect('db.sqlite3') as conn:
                conn.execute(
                    "INSERT INTO usuarios (email, nome, telefone, senha, tipo, ativo, data_expiracao) VALUES (?,?, ?, ?, ?, ?, ?)",
                    (email, nome, telefone, senha_hash, 'normal', 1, validade)
                )
                conn.commit()
            enviar_email(email, senha_gerada)
            return redirect('/admin/usuarios')
        return render_template('register.html')

    @app.route('/admin/cadastrar-em-massa', methods=['GET', 'POST'])
    def cadastrar_em_massa():
        if session.get('tipo') != 'admin':
            return redirect('/login')

        if request.method == 'POST':
            arquivo = request.files.get('arquivo')
            if not arquivo or not arquivo.filename.endswith('.xlsx'):
                return "Arquivo inválido. Envie um .xlsx com colunas email e telefone."
            try:
                wb = load_workbook(arquivo)
                ws = wb.active
            except Exception as e:
                return f"Erro ao abrir o Excel: {e}"

            usuarios_adicionados = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                email = row[0]
                telefone = row[1]
                nome = row[3]
                if not email or not telefone or not nome:
                    continue

                senha = 'ENES2025'
                senha_hash = generate_password_hash(senha)
                validade = (datetime.now() + timedelta(days=365)).strftime('%Y-%m-%d')
                try:
                    with sqlite3.connect('db.sqlite3') as conn:
                        conn.execute(
                            "INSERT INTO usuarios (email, nome, telefone, senha, tipo, ativo, data_expiracao) VALUES (?, ?, ?, ?, ?, ?, ? )",
                            (email, nome, telefone, senha_hash, 'normal', 1, validade)
                        )
                        conn.commit()
                    usuarios_adicionados.append(email)
                except sqlite3.IntegrityError:
                    continue

            return f"{len(usuarios_adicionados)} usuários cadastrados com sucesso."
        return render_template('cadastrar_em_massa.html')

    @app.route('/admin/usuarios')
    def listar_usuarios():
        if session.get('tipo') != 'admin':
            return redirect('/login')

        filtro = request.args.get('filtro', 'todos')  # ativos, inativos, todos
        query = """
            SELECT u.id, u.email, u.nome, u.tipo, u.ativo, u.data_expiracao, COUNT(r.id)
            FROM usuarios u
            LEFT JOIN formularios f ON f.user_id = u.id
            LEFT JOIN respostas r ON r.formulario_id = f.id
        """
        if filtro == 'ativos':
            query += " WHERE u.ativo=1"
        elif filtro == 'inativos':
            query += " WHERE u.ativo=0"
        query += " GROUP BY u.id ORDER BY u.email"

        with sqlite3.connect('db.sqlite3') as conn:
            usuarios = conn.execute(query).fetchall()
            
            # ---- NOVA QUERY para contar ativos/inativos ----
            status = conn.execute("""
                SELECT 
                    SUM(CASE WHEN ativo = 1 THEN 1 ELSE 0 END) as ativos,
                    SUM(CASE WHEN ativo = 0 THEN 1 ELSE 0 END) as inativos
                FROM usuarios
            """).fetchone()

        return render_template(
            'usuarios_admin.html',
            usuarios=usuarios,
            filtro=filtro,
            ativos=status[0] or 0,
            inativos=status[1] or 0
        )
    
    @app.route('/admin/usuario/<int:user_id>/toggle')
    def toggle_usuario(user_id):
        if session.get('tipo') != 'admin':
            return redirect('/login')

        with sqlite3.connect('db.sqlite3') as conn:
            status_atual = conn.execute("SELECT ativo FROM usuarios WHERE id=?", (user_id,)).fetchone()
            novo_status = 0 if status_atual[0] == 1 else 1
            conn.execute("UPDATE usuarios SET ativo=? WHERE id=?", (novo_status, user_id))
            conn.commit()

        return redirect('/admin/usuarios')
